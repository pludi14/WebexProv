from openpyxl import load_workbook
import pathlib
import logging
from log.setup_logger import logger



class Excelhandler():

    def __init__(self,orginfo, datei="" ):
        self.__dateiname = datei
        self.__path = pathlib.Path(__file__).parent.absolute()
        self.__workdir = pathlib.Path().absolute()
        self.__daten_sheet=""
        self.__anzahlDatensaetze=0
        self.__anzahl_Rows=0
        self.__OrgInfo=orginfo

        # Definiert in welcher Spalte der Excel die einzelnen Werte stehen.
        self.__col_Vorname=1
        self.__col_Nachname = 0
        self.__col_Email = 2
        self.__col_Messaging_Lic = 3
        self.__col_Meeting_Lic = 4
        self.__col_Kalender_Lic = 5
        self.__col_Call_Lic = 6
        #Definiert das Tabellenblatt
        self.__tabellenblatt="Daten"
        self.__startrow=2
        self.__anzahl_Cols=6

    logger = logging.getLogger("WP.excel")

    def __setDateiname(self,x): self.__dateiname =x
    def __getDateiname(self): return self.__dateiname
    dateiname=property(__getDateiname,__setDateiname)  # Get und set kan dann über den Variablennamen direkt ausgeführt werden.
                                                       # Nur für den Verwender der Klasse nicht innerhalb der Klasse  Verwenden

    def __setOrg_Info(self,x): self.__OrgInfo =x
    def __getOrg_Info(self): return self.__OrgInfo
    org_Info=property(__getOrg_Info,__setOrg_Info)


    def __get_Anzahl_Datensaetze(self): return self.__anzahlDatensaetze
    anzahl_Datensaetze=property(__get_Anzahl_Datensaetze)


    #Liest die Exce Datei ein.
    def leseExcel(self, datei=""):
        self.__dateiname = self.__dateiname if self.__dateiname else datei
        wb = load_workbook(filename=self.__dateiname, data_only=True) #Lade Excel. Data Only=Liest Values und nicht die Formeln als Wert aus.
        self.__daten_sheet=wb[self.__tabellenblatt]
        self.__anzahlDatensaetze=self.__getAnzahlDatensaetze()
        self.__anzahl_Rows=0
        for row in self.__daten_sheet.rows:
            self.__anzahl_Rows+=1


    def __getAnzahlDatensaetze(self):
        wertvorhanden=True
        zaehler=0
        row=2
        while wertvorhanden:
            zellenwert = self.__daten_sheet["C"+str(row)].value
            if zellenwert == None:
                wertvorhanden=False
            else:
                zaehler=zaehler+1
                row=row+1
        return zaehler

    # gibt Daten im gefordeten JSON Format zurück
    def getDaten(self):

        daten=[]
        lics={}


        for zeile in self.__daten_sheet.iter_rows(min_row=self.__startrow, max_col=self.__anzahl_Cols, max_row=self.__anzahl_Rows):
            lics["messaging"] = False
            lics["meeting"] = False
            lics["kalender"] = False
            lics["call"] = False
            if zeile[self.__col_Email].value:
                datensatz = {}
                datensatz["licenses"] = []
                datensatz["firstName"] = zeile[self.__col_Vorname].value
                datensatz["lastName"] = zeile[self.__col_Nachname].value
                datensatz["displayName"] = zeile[self.__col_Vorname].value+" "+zeile[self.__col_Nachname].value
                datensatz["emails"] = [zeile[self.__col_Email].value]

                if self.__checkUserinOrg(datensatz["emails"][0]):
                    datensatz["doing"]="update"
                else:
                    datensatz["doing"]="insert"

                if zeile[self.__col_Messaging_Lic].value == "x": lics["messaging"]=True
                if zeile[self.__col_Meeting_Lic].value == "x": lics["meeting"] = True
                if zeile[self.__col_Kalender_Lic].value == "x": lics["kalender"] = True
                lics["call"] = True
                datensatz["licenses"] = self.__set_correct_licenses(datensatz["emails"][0],lics)

                daten.append(datensatz)
        return daten
    # Prüft ob dier User in der ausgewählten Org ist
    def __checkUserinOrg(self, usermail):
        da_nichtda=False

        if usermail in self.__OrgInfo.org_Users:
            da_nichtda = True
        return da_nichtda

    # Setzt die korreketen LizenzIDs für den Benuterz
    def __set_correct_licenses(self,usermail,lics):
        messaging_LicID = self.__OrgInfo.messaging_lic_ID
        meeting_LicID = self.__OrgInfo.meeting_lic_ID
        h_calendar_LicID= self.__OrgInfo.calendar_lic_ID
        h_call_licID= self.__OrgInfo.call_lic_ID

        lics_of_user = []

        if self.__checkUserinOrg(usermail):
            lics_of_user = self.__OrgInfo.org_Users[usermail]["licenses"]

        for lic,true_or_false in lics.items():

            if lic == "messaging":
                if true_or_false:
                    if messaging_LicID not in lics_of_user: lics_of_user.append(messaging_LicID)
                else:
                    if messaging_LicID in lics_of_user: lics_of_user.remove(messaging_LicID)

            if lic == "meeting":
                if true_or_false:
                    if meeting_LicID not in lics_of_user: lics_of_user.append(meeting_LicID)
                else:
                    if meeting_LicID in lics_of_user: lics_of_user.remove(meeting_LicID)

            if lic == "kalender":
                if true_or_false:
                    if h_calendar_LicID not in lics_of_user: lics_of_user.append(h_calendar_LicID)
                else:
                    if h_calendar_LicID in lics_of_user: lics_of_user.remove(h_calendar_LicID)

            if lic == "call":
                if true_or_false:
                    if h_call_licID not in lics_of_user: lics_of_user.append(h_call_licID)
                else:
                    if h_call_licID in lics_of_user: lics_of_user.remove(h_call_licID)

        return lics_of_user




